import tkinter as tk
from tkinter import ttk
import os
import openpyxl
import datetime
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2TkAgg
from matplotlib.figure import Figure
import matplotlib.pyplot as plt

LARGE_FONT = ("微軟正黑體", 14)
SMALL_FONT = ("微軟正黑體", 10)

datalist = ['日期','收支','類別','金額','餘額',0]
expanse = ['飲食','交通','其他支出']
income = ['薪水','兼職','其他收入']

if os.path.exists('記帳.xlsx') != True:
		
			wb = openpyxl.Workbook()
			ws = wb['Sheet']
			ws.title ='明細資料'
			ws['A1'] = datalist[0]
			ws['B1'] = datalist[1]
			ws['C1'] = datalist[2]
			ws['D1'] = datalist[3]
			ws['E1'] = datalist[4]
			ws['F1'] = datalist[5]
			
			wb.create_sheet(title='支出分類')
			ws2 = wb['支出分類']
			for i in range(len(expanse)):
				ws2['A'+str(i)] = expanse[i]

			wb.create_sheet(title='收入分類')
			ws3 = wb['收入分類']
			for i in range(len(income)):
				ws3['A'+str(i)] = income[i]
			wb.save('記帳.xlsx')

wb = openpyxl.load_workbook('記帳.xlsx')

class App(tk.Tk):

	def __init__(self, *arg, **kwargs):
		
		tk.Tk.__init__(self, *arg, **kwargs)
		tk.Tk.wm_title(self, "記帳小幫手><")

		container = tk.Frame(self)
		container.pack(side="top", fill="both", expand = True)
		'''#抄來的code原本有，但沒有好像也沒差
		container.grid_rowconfigure(0, weight=1)
		container.grid_columnconfigure(0, weight=1)
		'''
		self.frames = {}

		for F in (StartPage, RecordPage, PieChart, BarChart):
			frame = F(container, self)
			self.frames[F] = frame
			frame.grid(row=0, column=0, sticky="nsew")
		
		self.show_frame(StartPage)

	def show_frame(self, cont):
		frame = self.frames[cont]
		frame.tkraise()

class StartPage(tk.Frame):
	
	def __init__(self, parent, controller):
		tk.Frame.__init__(self, parent)
		label = tk.Label(self, text="記帳小幫手><", font=LARGE_FONT)
		label2 = tk.Label(self, text="\n統計報表：")
		button = ttk.Button(self, text="記一筆",
			command=lambda: controller.show_frame(RecordPage))
		button2 = ttk.Button(self, text="分類圓餅圖",
			command=lambda: controller.show_frame(PieChart))
		button3 = ttk.Button(self, text="月長條圖",
			command=lambda: controller.show_frame(BarChart))
		label.pack(pady=10,padx=10)
		button.pack()
		label2.pack()
		button2.pack()
		button3.pack()

class RecordPage(tk.Frame):
	
	def __init__(self, parent, controller):
		tk.Frame.__init__(self, parent)
		label = tk.Label(self, text="新增收支紀錄", font=LARGE_FONT)
		label.grid(row=0,column=1,pady=20)

		label2 = tk.Label(self, text="日期：", font=SMALL_FONT)
		label2.grid(row=1,column=0,sticky="e")
		entry_date = tk.Entry(self)
		now = datetime.datetime.now()
		entry_date.insert(0,'{}/{}/{}'.format(now.year,now.month,now.day))
		entry_date.grid(row=1,column=1,columnspan=2,sticky="w")

		label3 = tk.Label(self, text="收支：", font=SMALL_FONT)
		label3.grid(row=2,column=0,sticky="e")
		combobox = ttk.Combobox(self, state="readonly",value=['支出','收入'])
		combobox.current(0)
		combobox.grid(row=2,column=1,columnspan=2,sticky="w")
		combobox.bind('<<ComboboxSelected>>', lambda x: set_type())

		label4 = tk.Label(self, text="類別：", font=SMALL_FONT)
		label4.grid(row=3,column=0,sticky="e")
		combobox2 = ttk.Combobox(self, state="readonly")
		combobox2['value']=expanse
		combobox2.current(0)
		combobox2.grid(row=3,column=1,columnspan=2,sticky="w")

		label6 = tk.Label(self, text="金額：", font=SMALL_FONT)
		label6.grid(row=4,column=0,sticky="e")
		entry = tk.Entry(self)
		entry.grid(row=4,column=1,columnspan=2,sticky="w")
		entry.insert(0,0)

		button3 = ttk.Button(self, text="<返回",
			command=lambda: controller.show_frame(StartPage))
		button3.grid(row=5,column=0,pady=15)
		button4 = ttk.Button(self, text="再記一筆",
			command=lambda: record())
		button4.grid(row=5,column=1)
		button5 = ttk.Button(self, text="儲存",
			command=lambda: [controller.show_frame(StartPage),record()])
		button5.grid(row=5,column=2)

		def set_type():
			nonlocal combobox
			nonlocal combobox2
			if combobox.get()=='支出':
				combobox2['value']= expanse
				combobox2.current(0)
			else :
				combobox2['value']= income
				combobox2.current(0)

		def record():
			ws = wb['明細資料']
			row = ws.max_row
			ws['A'+str(row+1)]=entry_date.get()
			ws['B'+str(row+1)]=combobox.get()
			ws['C'+str(row+1)]=combobox2.get()
			if combobox.get()=='收入':ws['D'+str(row+1)]=entry.get()
			else:ws['D'+str(row+1)]=str(-1*int(entry.get()))
			ws['F'+str(row+1)]=str(int(ws['F'+str(row)].value)+int(ws['D'+str(row+1)].value))

			if(combobox.get()=='收入'):
				ws2 = wb['收入分類']
				for t in range(income):
					if(combobox2.get()==t):
						ws2['B'+str(t)]+=entry.get()
			else:
				ws2 = wb['支出分類']
				for t in range(expanse):
					if(combobox2.get()==t):
						ws2['B'+str(t)]+=entry.get()
			wb.save('記帳.xlsx')
	
class PieChart(tk.Frame):

	def __init__(self, parent, controller):

		tk.Frame.__init__(self, parent)
		label = tk.Label(self, text="分類圓餅圖", font=LARGE_FONT)
		label.pack()
		combobox = ttk.Combobox(self, state="readonly",value=['支出','收入'])
		combobox.current(0)
		combobox.pack()
		combobox.bind('<<ComboboxSelected>>', lambda x: show_pie())
		ws = wb['支出分類']
		slices = []
		for row in range(ws.max_row):
			slices.append(ws['B'+str(row)])
		p = plt.pie(slices,labels=types, autopct='%1.1f%%')
		def show_pie():
			if (combobox.get() == '支出'):
				ws = wb['支出分類']
				for row in range(ws.max_row):
					slices.append(ws['B'+str(row)])
			else:
				ws = wb['收入分類']
				for row in range(ws.max_row):
					slices.append(ws['B'+str(row)])
		
		canvas = FigureCanvasTkAgg(p, self)
		canvas.draw()
		canvas.get_tk_widget().pack(side="top", fill="both", expand = True)

		button1 = ttk.Button(self, text="<返回",
			command=lambda: controller.show_frame(StartPage))
		button1.pack()
class BarChart(tk.Frame):

	def __init__(self, parent, controller):
		tk.Frame.__init__(self, parent)
		label = tk.Label(self, text="月長條圖", font=LARGE_FONT)
		label.pack()

		button1 = ttk.Button(self, text="<返回",
			command=lambda: controller.show_frame(StartPage))
		button1.pack()

		f = Figure(figsize=(5,5), dpi=50)
		a = f.add_subplot(111)
		a.plot([1,2,3,4,5,6,7,8,9,10,11,12],[5,6,1,3,8,9,3,5,5,5,5,5])

		canvas = FigureCanvasTkAgg(f, self)
		canvas.draw()
		canvas.get_tk_widget().pack(side="top", fill="both", expand = True)

app = App()
app.mainloop()
